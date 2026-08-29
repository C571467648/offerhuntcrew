# -*- coding: utf-8 -*-
"""
投递看板自动更新脚本（固定接口）
================================
用法：更新投递总表（Excel）后，运行本脚本即可自动重新生成 投递看板.html。

    python update_dashboard.py

数据源（固定）：
  05_选岗老师/数据/秋招央国企岗位投递表.xlsx    （sheet: 央国企岗位投递表）
  05_选岗老师/数据/秋招私企投递记录表.xlsx      （sheet: 遍历全部 sheet，按用户求职方向划分）

列名约定（勿改表头）：
  央国企：公司、网申日期、地点、岗位、是否投递、简历、笔试、一面、二面、三面、offer、链接、优先级、批次
  私企  ：公司、批次、测评截止时间、地点、优先级、岗位、是否投递、简历、测评、笔试、一面、二面、三面、offer、链接

状态推导：按 offer→三面→二面→一面→笔试→测评→简历 取最深非空阶段；
          阶段值：是→阶段名；通过→阶段+过；挂→阶段+挂（任何阶段"挂"=终态）。
投递明细：是否投递=是（投递表只记录已正式投递的岗位，计划投递统一放新增岗位情报）。
行动清单：'近期开闸·行动清单' = 新增岗位情报(央国企)中已开放网申的岗位（截止窗口含'网申中'），标注截止日期。
计划投递：'新增岗位情报' = 全部未投递的适合岗位（央国企/私企），按匹配度排序、持续累积、不重复不删除（已投递则移入投递明细）。

输出：05_选岗老师/可视化/投递看板.html（覆盖）
"""
import os
import re
import json
import datetime

# 用户姓名（看板标题显示）。按提示词指引由总架构师替换为使用者姓名，如 USER_NAME = '张同学'。
USER_NAME = '{{用户姓名}}'

BASE = os.path.dirname(os.path.abspath(__file__))          # 05_选岗老师/可视化
DATA_DIR = os.path.normpath(os.path.join(BASE, '..', '数据'))  # 05_选岗老师/数据
TIP_DIR = os.path.normpath(os.path.join(BASE, '..', '岗位情报'))  # 05_选岗老师/岗位情报
TEMPLATE = os.path.join(BASE, 'dashboard_template.html')
OUT = os.path.join(BASE, '投递看板.html')

SOE_FILE = '秋招央国企岗位投递表.xlsx'
PRV_FILE = '秋招私企投递记录表.xlsx'
TIPS_FILE = '新增岗位情报.xlsx'


def norm_cell(v):
    if v is None:
        return ''
    s = re.sub(r'\s+', ' ', str(v)).replace('\\n', ' ').strip()
    return '' if s.lower() == 'none' else s


def combo(stage, val):
    """把阶段列的值拼成状态串，如 简历+过 -> 简历过；笔试+待参加 -> 笔试待参加"""
    v = norm_cell(val)
    if not v:
        return None
    if v == '是':
        return stage
    if v == '通过':
        v = '过'
    return stage + v


def normalize_status(st):
    """把原始状态值归一化为筛选器使用的标准状态名（只允许 8 种）"""
    mapping = {
        '流程中': '简历筛选中',
        '简历过': '测评中',
        '简历筛选中': '简历筛选中',
        '简历挂': '简历挂',
        '测评': '测评中',
        '测评中': '测评中',
        '笔试已完成': '笔试中',
        '笔试中': '笔试中',
        '笔试挂': '笔试挂',
        '一面': '面试中',
        '二面': '面试中',
        '三面': '面试中',
        'offer': 'offer',
    }
    return mapping.get(st, '简历筛选中')  # 未知状态默认归为简历筛选中


def soe_status(row):
    for stage in ('offer', '三面', '二面', '一面', '笔试', '简历'):
        s = combo(stage, row.get(stage))
        if s:
            return normalize_status(s)
    return normalize_status('流程中')


def prv_status(row):
    for stage in ('offer', '三面', '二面', '一面', '笔试', '测评', '简历'):
        s = combo(stage, row.get(stage))
        if s:
            return normalize_status(s)
    return normalize_status('流程中')


def window_of(d):
    s = norm_cell(d)
    if '春招' in s:
        return '3-4月(春招)'
    if '8-9' in s or ('8' in s and '9' in s and '10' not in s and '11' not in s):
        return '8-9月'
    if '9-10' in s or '10' in s and '9' in s:
        return '9-10月'
    if '11-12' in s:
        return '11-12月'
    if '10-11' in s:
        return '10-11月'
    if '11' in s:
        return '11月'
    if '10' in s:
        return '10月'
    if '9' in s:
        return '9月'
    if '8' in s:
        return '8月'
    return '未标注'


BATCH_NORM = {'实习(练手)': '实习', '实习批次': '实习', '正式批次': '正式批'}


def norm_batch(b):
    v = norm_cell(b)
    if v.startswith('实习'):
        return '实习'
    return BATCH_NORM.get(v, v) or '-'


def load_soe(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], [], []
    header = list(rows[0])
    applied, pool, win_cnt = [], [], {}
    for r in rows[1:]:
        if not any(v is not None and str(v).strip() for v in r):
            continue
        d = dict(zip(header, r))
        co = norm_cell(d.get('公司'))
        if not co:
            continue
        rec = {
            'co': co,
            'date': norm_cell(d.get('网申日期')),
            'loc': norm_cell(d.get('地点')),
            'job': norm_cell(d.get('岗位')),
            'link': norm_cell(d.get('链接')),
            'pri': norm_cell(d.get('优先级')),
            'batch': norm_cell(d.get('批次')),
        }
        if norm_cell(d.get('是否投递')) == '是':
            rec['st'] = soe_status(d)
            applied.append(rec)
        else:
            rec['win'] = window_of(d.get('网申日期'))
            pool.append(rec)
            win_cnt[rec['win']] = win_cnt.get(rec['win'], 0) + 1
    # 窗口按固定顺序输出，未标注放最后
    order = ['8-9月', '9月', '9-10月', '10月', '10-11月', '11月', '11-12月', '3-4月(春招)', '未标注']
    keys = sorted(win_cnt, key=lambda k: order.index(k) if k in order else 99)
    windows = [{'name': k, 'n': win_cnt[k]} for k in keys]
    return applied, pool, windows


def load_prv(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out, pool = [], []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = list(rows[0])
        for r in rows[1:]:
            if not any(v is not None and str(v).strip() for v in r):
                continue
            d = dict(zip(header, r))
            co = norm_cell(d.get('公司'))
            if not co:
                continue
            rec = {
                'dir': sheet,
                'co': co,
                'batch': norm_batch(d.get('批次')),
                'loc': norm_cell(d.get('地点')),
                'pri': norm_cell(d.get('优先级')) or '-',
                'job': norm_cell(d.get('岗位')),
                'link': norm_cell(d.get('链接')),
            }
            if norm_cell(d.get('是否投递')) == '是':
                rec['st'] = prv_status(d)
                out.append(rec)
            else:
                pool.append(rec)
    wb.close()
    return out, pool


TIPS_KEYS = ('公司', '性质', '届别', '岗位', '地点', '截止窗口', '匹配度', '备注', '入口')


def is_open(win):
    """判定岗位是否已开放网申。

    规则（用户 2026-08-29 确认）：
      - '截止窗口'含'网申中' → 已开放
      - 含'未开放'或其他明确未开放描述 → 未开放
      - 空/无法判定 → 视为已开放（宁可出错不可遗漏）
    由选岗老师维护时把'截止窗口'列写规范（'网申中·…'/'未开放·…'）。
    """
    s = norm_cell(win)
    if not s:
        return True
    return '网申中' in s


def load_jobtips(path):
    """读 新增岗位情报.xlsx：分类列=央国企→soe 模块，私企→prv 模块；文件缺失时返回空（兼容开源版初始状态）。
    每条记录附 open 标记（是否已开放网申），供'近期开闸·行动清单'使用。"""
    if not os.path.exists(path):
        return [], []
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = list(rows[0])
    soe, prv = [], []
    for r in rows[1:]:
        if not any(v is not None and str(v).strip() for v in r):
            continue
        d = dict(zip(header, r))
        co = norm_cell(d.get('公司'))
        if not co:
            continue
        rec = {k: norm_cell(d.get(k)) for k in TIPS_KEYS}
        rec['open'] = is_open(rec['截止窗口'])
        (soe if norm_cell(d.get('分类')) == '央国企' else prv).append(rec)
    return soe, prv


def tips_date():
    """从最新情报清单文件名提取日期（如 2026-08-27_...xlsx → '8-27'），无则返回空串。"""
    if not os.path.isdir(TIP_DIR):
        return ''
    files = [f for f in os.listdir(TIP_DIR) if f.endswith('.xlsx') and '情报清单' in f]
    if not files:
        return ''
    files.sort(reverse=True)
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', files[0])
    return '%s-%s' % (m.group(2).lstrip('0'), m.group(3)) if m else ''


def main():
    soe_applied, soe_pool, soe_windows = load_soe(os.path.join(DATA_DIR, SOE_FILE))
    prv, prv_pool = load_prv(os.path.join(DATA_DIR, PRV_FILE))
    soe_tips, prv_tips = load_jobtips(os.path.join(TIP_DIR, TIPS_FILE))

    # 近期开闸·行动清单 = 新增岗位情报(央国企)中已开放网申的岗位，标注截止窗口
    soe_act = [{
        'co': r['公司'],
        'deadline': r['截止窗口'],
        'loc': r['地点'],
        'job': r['岗位'],
        'link': r['入口'],
    } for r in soe_tips if r['open']]

    data = {
        'date': datetime.date.today().strftime('%Y-%m-%d'),
        'soe': {'applied': soe_applied, 'pool': soe_pool, 'windows': soe_windows},
        'prv': prv,
        'prv_pool': prv_pool,
        'soe_tips': soe_tips,
        'prv_tips': prv_tips,
        'soe_act': soe_act,
        'tips_date': tips_date(),
    }
    payload = json.dumps(data, ensure_ascii=False, indent=1).replace('</', '<\\/')

    with open(TEMPLATE, 'r', encoding='utf-8') as f:
        html = f.read()
    html = html.replace('/*__DATA__*/ null', '/*__DATA__*/ ' + payload)
    html = html.replace('{{USER_NAME}}', USER_NAME)

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(html)

    total_prv = len(prv)
    dead_prv = sum(1 for r in prv if r['st'].endswith('挂'))
    print('[OK] 投递看板已生成: %s' % OUT)
    print('     央国企: 已投 %d 家, 新增岗位情报(计划投递) %d 家, 近期开闸行动清单 %d 家' % (len(soe_applied), len(soe_tips), len(soe_act)))
    print('     私企  : 共 %d 条, 已挂 %d 条' % (total_prv, dead_prv))
    print('     新增岗位情报: 央国企 %d 条, 私企 %d 条' % (len(soe_tips), len(prv_tips)))


if __name__ == '__main__':
    main()
